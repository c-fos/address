#!/usr/bin/python3
# coding=utf-8
"""
Модуль для извлечения информации из текстового и табличного представлений профиля настроек
"""

# Импорт необходимых модулей

import pprint
import logging
import argparse

from openpyxl import load_workbook
import xlsxwriter

logger = logging.getLogger()
logging.basicConfig(filename='test.log', filemode='w', level=logging.DEBUG)


def compare_sections(main_dict, new_dict):
    for section, s_value in new_dict['sections'].items():
        if section in main_dict['sections']:
            if len(new_dict['sections']) > 1:
                # Вариант 3
                logger.debug('Пошли по ветке 3')
                if section is None:
                    if 'бн' in main_dict['sections']:
                        logger.debug('Добавили счет в секцию бн')
                        add_accounts(main_dict['sections']['бн'], s_value)
                    else:
                        logger.debug('Добавили секцию бн')
                        add_section(main_dict, 'бн', s_value)
                else:
                    logger.debug('Добавили счет в секцию бн')
                    add_accounts(main_dict['sections'][section], s_value)
            else:
                # варианты 1, 4
                logger.debug('Пошли по ветке 1,4')
                add_accounts(main_dict['sections'][section], s_value)
        else:
            if section is None and 'бн' not in main_dict['sections']:
                # вариант 2
                logger.debug('Пошли по ветке 2. добавили секцию бн')
                add_section(main_dict, 'бн', s_value)
            elif section is None and 'бн' in main_dict['sections']:
                # вариант 2
                logger.debug('Пошли по ветке 2. добавили в секцию бн')
                add_accounts(main_dict['sections']['бн'], s_value)
            else:
                # вариант 4
                logger.debug('Пошли по ветке 4')
                add_section(main_dict, section, s_value)


def add_accounts(main_section, new_section):
    for acc, a_value in new_section['accounts'].items():
        assert acc not in main_section['accounts']
        main_section['accounts'][acc] = a_value


def add_section(main_dict, section_name, new_section):
    main_dict['sections'][section_name] = new_section


def area_recalc(section):
    area = 0
    for acc, a_value in section['accounts'].items():
        area += a_value['owarea']
    section['meta']['full_area'] = area


def get_prem_type(main_dict, addr):
    if None in main_dict['sections'] and len(main_dict['sections']) == 1:
        new_prem_type = 1
    elif None in main_dict['sections'] and len(main_dict['sections']) > 1:
        if 'бн' in main_dict['sections']:
            add_accounts(main_dict['sections']['бн'], main_dict['sections'][None])
        else:
            main_dict['sections']['бн'] = main_dict['sections'][None]
        del main_dict['sections'][None]
        new_prem_type = 2
    else:
        new_prem_type = 2
    if new_prem_type != main_dict['meta']['prem_type']:
        logger.warning('У адреса {} изменет prem_type на {}'.format(addr, new_prem_type))
        main_dict['meta']['prem_type'] = new_prem_type


def postprocessing(data_str):
    for addr, a_value in data_str.items():
        get_prem_type(a_value, addr)
        for section, s_value in a_value['sections'].items():
            area_recalc(s_value)


def write_xlsx(data_str, filename):

    """
    Функция в которой на основании переданного основного словаря формируется
    xlsx таблица.

    """
    workbook = xlsxwriter.Workbook(filename, {'constant_memory': True})
    worksheet = workbook.add_worksheet()
    columns = ['AREA', 'DISTRICT', 'CITY', 'STREET', 'BUILDING', 'BILDBULK', 'BSECTION', 'FLAT', 'FSECTION', 'CONTRNUM',
               'FULLAREA', 'PREMTYPE', 'OWTYPE', 'OWAREA']
    row_number = 0
    for ind, name in enumerate(columns):
        worksheet.write(row_number, ind, name)
    for addr, a_value in data_str.items():
        for section, s_value in a_value['sections'].items():
            for acc, acc_value in s_value['accounts'].items():
                row_number += 1
                tmp_data = {'AREA': addr[0],
                            'DISTRICT': addr[1],
                            'CITY': addr[2],
                            'STREET': addr[3],
                            'BUILDING': addr[4],
                            'BILDBULK': addr[5],
                            'BSECTION': addr[6],
                            'FLAT': addr[7],
                            'FSECTION': section,
                            'CONTRNUM': acc,
                            'FULLAREA': s_value['meta']['full_area'],
                            'PREMTYPE': a_value['meta']['prem_type'],
                            'OWTYPE': 'AREA',
                            'OWAREA': acc_value['owarea']}
                for ind, col_name in enumerate(columns):
                    worksheet.write(row_number, ind, tmp_data[col_name])
    workbook.close()


def main_dict_from_xlsx(filename):
    """
    """
    wb = load_workbook(filename=filename, read_only=True)
    ws = wb.active
    rows = ws.rows
    # чтение и разбор строки
    # чтения названий колонок
    name_row = next(rows)
    col_name_to_ind = {j: i for i, j in enumerate(map(lambda x: str(x.value), name_row))}
    logger.info(col_name_to_ind)
    addr_names = ['AREA', 'DISTRICT', 'CITY', 'STREET', 'BUILDING', 'BILDBULK', 'BSECTION', 'FLAT']
    section_name = 'FSECTION'
    account_name = 'CONTRNUM'
    data_str = {}
    # чтение данных
    for row_index, row in enumerate(rows):
        try:
            # address: AREA	DISTRICT	CITY	STREET	BUILDING	BILDBULK	BSECTION	FLAT
            tmp_addr = (row[col_name_to_ind[i]].value for i in addr_names)
            addr = tuple(map(lambda x: str(x) if x else None, tmp_addr))
            section = row[col_name_to_ind[section_name]].value
            section = str(section) if section else None
            account = row[col_name_to_ind[account_name]].value
            area = float(row[col_name_to_ind['OWAREA']].value)
            section_area = float(row[col_name_to_ind['FULLAREA']].value) if 'FULLAREA' in col_name_to_ind else None
            prem_type = row[col_name_to_ind['PREMTYPE']].value if 'PREMTYPE' in col_name_to_ind else None
            if addr in data_str:
                assert(data_str[addr]['meta']['prem_type'] == prem_type)
                if section in data_str[addr]['sections']:
                    assert(data_str[addr]['sections'][section]['meta']['full_area'] == section_area)
                    if account in data_str[addr]['sections'][section]['accounts']:
                        logger.debug("Внимание уже есть запись с совпадающими адресом,"
                                     " секцией и счетом. строка {}".format(row_index + 2))
                    else:
                        data_str[addr]['sections'][section]['accounts'][account] = {'owarea': area}
                else:
                    data_str[addr]['sections'][section] = {'meta': {'full_area': section_area},
                                                           'accounts': {account: {'owarea': area}}
                                                           }
            else:
                data_str[addr] = {'meta': {'prem_type': prem_type},
                                  'sections': {section: {'meta': {'full_area': section_area},
                                                         'accounts': {account: {'owarea': area}}
                                                         }
                                               }
                                  }
        except AssertionError:
            logger.debug("Ошибка при обработке строки %s!", row_index + 2)
    return data_str


def run(options):
    debug = options.debug
    data_str_full = main_dict_from_xlsx(options.main_file)
    data_str_new = main_dict_from_xlsx(options.new_file)
    if debug:
        with open('main_dict_before.txt', 'w') as fd:
            pprint.pprint(data_str_full, stream=fd)
        with open('new_dict.txt', 'w') as fd:
            pprint.pprint(data_str_new, stream=fd)
    counter = 0
    for key, val in data_str_new.items():
        if key in data_str_full:
            counter += 1
            logger.debug('Совпадения адреса: {}'.format(key))
            compare_sections(data_str_full[key], val)
        else:
            logger.debug('Адрес {} добавлен в основной файл'.format(key))
            data_str_full[key] = val
    postprocessing(data_str_full)
    logger.debug('Количество адресов в основном файле: {}'.format(len(data_str_full)))
    logger.debug('Количество адресов в добавляемом файле: {}'.format(len(data_str_new)))
    logger.debug("Количество совпадений: {}".format(counter))
    if debug:
        with open('main_dict_after.txt', 'w') as fd:
            pprint.pprint(data_str_full, stream=fd)
    write_xlsx(data_str_full, filename='result.xlsx')

if __name__ == "__main__":
    def cli():
        """
        Интерфейс командной строки

        Функция обрабатывает введенные в командной строке флаги, выводит help,
        возвращает имя файла с настройками
        """

        parser = argparse.ArgumentParser(description="Программа для объеденения двух xlsx файлов с данными по адресам")
        parser.add_argument("main_file", help="Имя основного xlsx файла")
        parser.add_argument("new_file", help="Имя нового xlsx файла")
        parser.add_argument("-d", "--debug", dest="debug", help="Выводить отладочную информацию")
        _options = parser.parse_args()
        return _options
    run(cli())
